import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    
    # Fills
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    accent_green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    warning_yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEAD2", end_color="FFEAD2", fill_type="solid")
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    double_bottom_side = Side(border_style="double", color="1B365D")
    thin_top_side = Side(border_style="thin", color="1B365D")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=thin_top_side, bottom=double_bottom_side)

    # ----------------------------------------------------
    # SHEET 1: CONFIGURACIÓN
    # ----------------------------------------------------
    ws_config = wb.create_sheet(title="Configuración")
    ws_config.views.sheetView[0].showGridLines = True
    
    ws_config["A1"] = "PARÁMETROS DE CONFIGURACIÓN FINANCIERA Y TRIBUTARIA"
    ws_config["A1"].font = title_font
    ws_config["A2"] = "Variables globales de tasas horarias, impuestos (SUNAT Régimen MYPE), detracciones y pasarelas de pago."
    ws_config["A2"].font = subtitle_font
    
    # Table headers
    headers = ["Concepto / Variable", "Valor", "Unidad", "Descripción"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_config.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    config_data = [
        # Mano de Obra (Costo Interno)
        ("Tasa Horaria Alberto (Costo MOD)", 93.75, "PEN/hora", "Tu valor por hora base para estimación interna ($25.00 * 3.75)"),
        ("Tasa Horaria Alberto (Venta)", 150.00, "PEN/hora", "Tu valor por hora cobrado para modificaciones/consultoría ($40.00 * 3.75)"),
        ("Tasa Horaria Junior (Costo MOD)", 30.00, "PEN/hora", "Tarifa horaria para programador junior de apoyo ($8.00 * 3.75)"),
        ("Tasa Horaria QA / Soporte (Costo MOD)", 22.50, "PEN/hora", "Tarifa horaria para asistente de pruebas de calidad ($6.00 * 3.75)"),
        
        # Localización Perú & SUNAT
        ("Tipo de Cambio (TC)", 3.75, "Soles por USD", "Tipo de cambio oficial SUNAT promedio"),
        ("UIT Vigente", 5150.00, "Soles (PEN)", "Valor de la Unidad Impositiva Tributaria para el año fiscal"),
        ("Tasa de Detracción (Software)", 0.12, "Porcentaje", "Detracción obligatoria para servicios de software > S/. 700 (12%) - Código 022"),
        ("Tasa de IGV", 0.18, "Porcentaje", "Impuesto General a las Ventas aplicable en Perú (18%)"),
        ("RMT Pago a Cuenta Mensual", 0.01, "Porcentaje", "Régimen MYPE Tributario - Pago a cuenta de Renta mensual (1.0% para < 300 UIT)"),
        ("RMT Provisión Renta Anual", 0.10, "Porcentaje", "Tasa de regularización anual del RMT (10% hasta 15 UIT de utilidad)"),
        
        # Pasarelas e Intermediación
        ("PayPal Comisión Fija", 0.30, "USD", "Tarifa transaccional fija de PayPal por pago internacional"),
        ("PayPal Comisión Porcentual", 0.054, "Porcentaje", "Comisión porcentual de PayPal por recepción internacional (5.4%)"),
        ("Wise Comisión Porcentual", 0.002, "Porcentaje", "Comisión porcentual por envío mediante Wise (0.2%)"),
        
        # Parámetros del Cotizador
        ("Buffer de Contingencia (PMBOK)", 0.20, "Porcentaje", "Reserva de contingencia para imprevistos técnicos (20%)"),
        ("Margen Comercial Objetivo", 0.40, "Porcentaje", "Margen comercial neto esperado sobre la venta sugerida (40%)")
    ]
    
    for row_idx, data in enumerate(config_data, 5):
        ws_config.cell(row=row_idx, column=1, value=data[0]).font = regular_font
        
        val_cell = ws_config.cell(row=row_idx, column=2, value=data[1])
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        
        # Formats
        if data[2] in ["PEN/hora", "Soles (PEN)"]:
            val_cell.number_format = 'S/. #,##0.00'
        elif data[2] in ["Soles por USD", "USD"]:
            val_cell.number_format = '$#,##0.00'
        elif data[2] == "Porcentaje":
            val_cell.number_format = '0.0%'
            
        ws_config.cell(row=row_idx, column=3, value=data[2]).font = regular_font
        ws_config.cell(row=row_idx, column=4, value=data[3]).font = regular_font
        
        for c in range(1, 5):
            ws_config.cell(row=row_idx, column=c).border = border_all
            ws_config.cell(row=row_idx, column=c).fill = light_blue_fill if row_idx % 2 == 0 else white_fill

    # ----------------------------------------------------
    # SHEET 2: OVERHEAD Y CAPACIDAD (TDABC)
    # ----------------------------------------------------
    ws_oh = wb.create_sheet(title="Overhead y Capacidad")
    ws_oh.views.sheetView[0].showGridLines = True
    
    ws_oh["A1"] = "CAPACIDAD PRODUCTIVA Y PRORRATEO DE GASTOS INDIRECTOS (TDABC)"
    ws_oh["A1"].font = title_font
    ws_oh["A2"] = "Cálculo de horas facturables reales y costo del overhead mensual amortizado por hora."
    ws_oh["A2"].font = subtitle_font
    
    # Capacidad laboral
    ws_oh["A4"] = "1. Capacidad Laboral Facturable (TDABC)"
    ws_oh["A4"].font = bold_font
    
    headers_cap = ["Métrica de Capacidad", "Valor", "Unidad", "Lógica / Referencia"]
    for col_idx, h in enumerate(headers_cap, 1):
        cell = ws_oh.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    cap_data = [
        ("Días del Año", 365, "Días", "Año natural"),
        ("Fines de Semana", 104, "Días", "52 fines de semana al año"),
        ("Feriados Nacionales (Perú)", 16, "Días", "Feriados oficiales calendario nacional"),
        ("Vacaciones / Descanso anual", 15, "Días", "Periodo de descanso proyectado para Alberto"),
        ("Días estimados de Enfermedad", 5, "Días", "Reserva de seguridad por imprevistos médicos"),
        ("Días Laborables Anuales", "=B6-B7-B8-B9-B10", "Días", "Días laborables reales disponibles"),
        ("Horas Diarias Laborables", 8, "Horas", "Jornada estándar"),
        ("Tasa de Utilización Facturable", 0.60, "Porcentaje", "Porcentaje de tiempo dedicado a producir código facturable (60%)"),
        ("Horas Facturables Anuales", "=B11*B12*B13", "Horas", "Horas útiles para proyectos facturables al año"),
        ("Horas Facturables Mensuales", "=B14/12", "Horas", "Capacidad mensual real facturable promedio (90 horas)")
    ]
    
    for row_idx, data in enumerate(cap_data, 6):
        ws_oh.cell(row=row_idx, column=1, value=data[0]).font = regular_font
        
        val_cell = ws_oh.cell(row=row_idx, column=2, value=data[1])
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        
        if data[2] == "Días" or data[2] == "Horas":
            val_cell.number_format = '#,##0'
        elif data[2] == "Porcentaje":
            val_cell.number_format = '0.0%'
            
        ws_oh.cell(row=row_idx, column=3, value=data[2]).font = regular_font
        ws_oh.cell(row=row_idx, column=4, value=data[3]).font = regular_font
        
        for c in range(1, 5):
            ws_oh.cell(row=row_idx, column=c).border = border_all
            ws_oh.cell(row=row_idx, column=c).fill = light_blue_fill if row_idx % 2 == 0 else white_fill
            
    # Estructura de Gastos Fijos (Overhead)
    ws_oh["A17"] = "2. Estructura de Gastos Fijos y Amortización Mensual (Soles PEN)"
    ws_oh["A17"].font = bold_font
    
    headers_exp = ["Concepto / Gasto Fijo", "Costo Mensual", "Unidad", "Detalle de Prorrateo / Causalidad SUNAT"]
    for col_idx, h in enumerate(headers_exp, 1):
        cell = ws_oh.cell(row=18, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    exp_data = [
        ("VPS de Desarrollo (Hetzner Cloud)", 225.00, "S/. / mes", "Servidor de staging e integraciones propias ($60.00 USD)"),
        ("ChatGPT Plus / Claude Pro (Suscripción IA)", 75.00, "S/. / mes", "Suscripción para asistencia técnica y código ($20.00 USD)"),
        ("Google Workspace & One (Almacenamiento)", 33.75, "S/. / mes", "Correo corporativo y almacenamiento Drive ($9.00 USD)"),
        ("Asesoría Contable / Contador local", 375.00, "S/. / mes", "Servicios contables mensuales para la declaración RUC 10"),
        ("Servicios del Hogar (Prorrateo 50% internet/luz)", 140.00, "S/. / mes", "Deducción de luz (S/. 40), agua (S/. 30), internet (S/. 70) por oficina en casa"),
        ("Depreciación Laptop de Desarrollo (25% SUNAT)", 75.00, "S/. / mes", "Laptop de S/. 3,600 depreciada linealmente a 4 años contables"),
        ("Fondo de Ahorro para Laptop (Financiero)", 75.00, "S/. / mes", "Diferencial de fondo líquido para renovar equipo en 2 años reales"),
        ("Total Overhead Mensual (Gastos Generales)", "=SUM(B19:B25)", "S/. / mes", "Total costo indirecto a absorber por tus proyectos"),
        ("Tasa de Absorción de Costos Indirectos (TACI)", "=B26/B15", "S/. por hora", "Costo indirecto por cada hora de desarrollo a facturar (Overhead/Horas)")
    ]
    
    for row_idx, data in enumerate(exp_data, 19):
        ws_oh.cell(row=row_idx, column=1, value=data[0]).font = regular_font
        
        val_cell = ws_oh.cell(row=row_idx, column=2, value=data[1])
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = 'S/. #,##0.00'
            
        ws_oh.cell(row=row_idx, column=3, value=data[2]).font = regular_font
        ws_oh.cell(row=row_idx, column=4, value=data[3]).font = regular_font
        
        for c in range(1, 5):
            ws_oh.cell(row=row_idx, column=c).border = border_all
            if data[0] in ["Total Overhead Mensual (Gastos Generales)", "Tasa de Absorción de Costos Indirectos (TACI)"]:
                ws_oh.cell(row=row_idx, column=c).fill = accent_green_fill
                ws_oh.cell(row=row_idx, column=c).font = bold_font
            else:
                ws_oh.cell(row=row_idx, column=c).fill = light_blue_fill if row_idx % 2 == 0 else white_fill

    # ----------------------------------------------------
    # SHEET 3: CALCULADORA INTERACTIVA (WBS)
    # ----------------------------------------------------
    ws_calc = wb.create_sheet(title="Calculadora WBS")
    ws_calc.views.sheetView[0].showGridLines = True
    
    ws_calc["A1"] = "COTIZADOR INTERACTIVO BASADO EN TAREAS (WBS / EDT)"
    ws_calc["A1"].font = title_font
    ws_calc["A2"] = "Escribe tus tareas de WBS, selecciona el rol (Alberto, Junior, QA) en la columna D y define las horas. Todo se calculará con fórmulas dinámicas."
    ws_calc["A2"].font = subtitle_font
    
    # Section Inputs
    ws_calc["A4"] = "1. ESTRUCTURA DE DESGLOSE DEL TRABAJO (WBS)"
    ws_calc["A4"].font = bold_font
    
    headers_wbs = [
        "ID", "Entregable / Fase", "Actividad del Software / IA", 
        "Rol Responsable", "Horas", "Costo MOD/h", "Total MOD", "Tasa TACI", "Overhead CIF"
    ]
    for col_idx, h in enumerate(headers_wbs, 1):
        cell = ws_calc.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    tasks_wbs = [
        ("1", "Levantamiento y Diseño", "Reuniones de arquitectura y mockup UI/UX inicial", "Alberto", 6),
        ("2", "Desarrollo Frontend SPA", "Maquetación de pantallas en React/TS", "Junior", 15),
        ("3", "Integración de Base de Datos", "Creación de esquemas Postgres/MongoDB", "Junior", 8),
        ("4", "Desarrollo Backend API", "Orquestación de lógica en Node.js/Python", "Alberto", 10),
        ("5", "Integración Silhouette Brain", "Configuración de enjambres en n8n e IA", "Alberto", 12),
        ("6", "Conexión WhatsApp API", "Vinculación de webhook Meta Cloud API", "Junior", 8),
        ("7", "Pruebas de Calidad (QA)", "Testing de estrés y verificación de prompts", "QA", 10),
        ("8", "Despliegue y CI/CD", "Dockerización, pipelines y hosting en VPS", "Alberto", 4),
        ("9", "Documentación Técnica", "Manual operativo y video Loom", "Junior", 4),
        ("10", "Soporte Inicial", "Atención a bugs durante la puesta en marcha", "QA", 5)
    ]
    
    for row_idx, t in enumerate(tasks_wbs, 6):
        ws_calc.cell(row=row_idx, column=1, value=t[0]).font = regular_font
        ws_calc.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        
        ws_calc.cell(row=row_idx, column=2, value=t[1]).font = regular_font
        ws_calc.cell(row=row_idx, column=3, value=t[2]).font = regular_font
        
        # Rol (Yellow fill for input)
        rol_cell = ws_calc.cell(row=row_idx, column=4, value=t[3])
        rol_cell.font = bold_font
        rol_cell.fill = warning_yellow_fill
        rol_cell.alignment = Alignment(horizontal="center")
        
        # Horas (Yellow fill for input)
        hrs_cell = ws_calc.cell(row=row_idx, column=5, value=t[4])
        hrs_cell.font = bold_font
        hrs_cell.fill = warning_yellow_fill
        hrs_cell.alignment = Alignment(horizontal="right")
        hrs_cell.number_format = '#,##0'
        
        # Tarifa MOD/h (Formula)
        r = row_idx
        mod_cell = ws_calc.cell(row=r, column=6, value=f"=IF(D{r}=\"Alberto\", Configuración!$B$5, IF(D{r}=\"Junior\", Configuración!$B$7, IF(D{r}=\"QA\", Configuración!$B$8, 0)))")
        mod_cell.font = regular_font
        mod_cell.alignment = Alignment(horizontal="right")
        mod_cell.number_format = 'S/. #,##0.00'
        
        # Total MOD (Formula)
        tot_mod_cell = ws_calc.cell(row=r, column=7, value=f"=E{r}*F{r}")
        tot_mod_cell.font = bold_font
        tot_mod_cell.alignment = Alignment(horizontal="right")
        tot_mod_cell.number_format = 'S/. #,##0.00'
        
        # Tasa TACI (Formula)
        taci_cell = ws_calc.cell(row=r, column=8, value="='Overhead y Capacidad'!$B$27")
        taci_cell.font = regular_font
        taci_cell.alignment = Alignment(horizontal="right")
        taci_cell.number_format = 'S/. #,##0.00'
        
        # Total Overhead (Formula)
        tot_oh_cell = ws_calc.cell(row=r, column=9, value=f"=E{r}*H{r}")
        tot_oh_cell.font = bold_font
        tot_oh_cell.alignment = Alignment(horizontal="right")
        tot_oh_cell.number_format = 'S/. #,##0.00'
        
        for c in range(1, 10):
            ws_calc.cell(row=r, column=c).border = border_all
            if not ws_calc.cell(row=r, column=c).fill.fill_type:
                ws_calc.cell(row=r, column=c).fill = light_blue_fill if r % 2 == 0 else white_fill

    # Section 2: Consolidado de Costos
    ws_calc["A17"] = "2. CONSOLIDADO DE COSTOS (Resumen Técnico)"
    ws_calc["A17"].font = bold_font
    
    headers_con = ["Elemento de Costo", "Valor Calculado", "Unidad", "Referencia / Lógica"]
    for col_idx, h in enumerate(headers_con, 1):
        cell = ws_calc.cell(row=18, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    con_data = [
        ("Total Horas de Ingeniería del WBS", "=SUM(E6:E15)", "Horas", "Esfuerzo neto estimado de todo el equipo"),
        ("Total Mano de Obra Directa (MOD)", "=SUM(G6:G15)", "Soles (PEN)", "Costo directo por desarrollo de software"),
        ("Total Costo Indirecto de Operación (CIF)", "=SUM(I6:I15)", "Soles (PEN)", "Financiamiento prorrateado del overhead mensual (TACI × horas)"),
        ("Materiales e Infraestructura Directos (MID)", 18.75, "Soles (PEN)", "Tokens de IA para pruebas y licencias directas (Celda amarilla editable)"),
        ("Costo Técnico de Producción", "=B20+B21+B22", "Soles (PEN)", "MOD + CIF + MID"),
        ("Reserva de Contingencia (20% PMBOK)", "=B23*Configuración!$B$18", "Soles (PEN)", "Fondo de contingencia para mitigar riesgos y alucinaciones"),
        ("Costo Unitario Total (Break-Even Real)", "=B23+B24", "Soles (PEN)", "Costo total de ruptura. Cobrar menos es pérdida líquida"),
        ("Precio Mínimo de Seguridad (Margen 20%)", "=B25/(1-0.20)", "Soles (PEN)", "Costo Unitario / 0.8 (Garantiza 20% utilidad neta)"),
        ("Precio de Venta Sugerido (Margen 40%)", "=B25/(1-Configuración!$B$19)", "Soles (PEN)", "Costo Unitario / 0.6 (Margen objetivo del 40%)")
    ]
    
    for idx, con in enumerate(con_data, 19):
        ws_calc.cell(row=idx, column=1, value=con[0]).font = regular_font
        
        val_cell = ws_calc.cell(row=idx, column=2, value=con[1])
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        
        if con[0] == "Total Horas de Ingeniería del WBS":
            val_cell.number_format = '#,##0'
        elif con[0] == "Materiales e Infraestructura Directos (MID)":
            val_cell.number_format = 'S/. #,##0.00'
            val_cell.fill = warning_yellow_fill
        else:
            val_cell.number_format = 'S/. #,##0.00'
            
        ws_calc.cell(row=idx, column=3, value=con[2]).font = regular_font
        ws_calc.cell(row=idx, column=4, value=con[3]).font = regular_font
        
        for c in range(1, 5):
            cell = ws_calc.cell(row=idx, column=c)
            cell.border = border_all
            if con[0] == "Costo Unitario Total (Break-Even Real)":
                cell.fill = light_blue_fill
                cell.font = bold_font
            elif con[0] in ["Precio Mínimo de Seguridad (Margen 20%)", "Precio de Venta Sugerido (Margen 40%)"]:
                cell.fill = accent_green_fill
                cell.font = bold_font

    # Section 3: Tributación y Liquidación
    ws_calc["A29"] = "3. LIQUIDACIÓN COMERCIAL, IMPUESTOS Y PASARELAS"
    ws_calc["A29"].font = bold_font
    
    for col_idx, h in enumerate(headers_con, 1):
        cell = ws_calc.cell(row=30, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    tax_data = [
        ("Ubicación del Cliente (PE = Perú / INT = Internacional)", "PE", "Código", "Celda amarilla editable (PE / INT)"),
        ("Canal de Pago (PP = PayPal / WB = Wise / TR = Local)", "PP", "Código", "Celda amarilla editable (PP / WB / TR)"),
        ("Precio Sugerido con Gross-Up", "=IF(B32=\"PP\", (B27+Configuración!$B$15*Configuración!$B$9)/(1-Configuración!$B$16), IF(B32=\"WB\", B27/(1-Configuración!$B$17), B27))", "Soles (PEN)", "Amortiza las comisiones financieras de retiro internacional"),
        ("Tasa de IGV", "=IF(B31=\"PE\", Configuración!$B$12, 0.0)", "Porcentaje", "18% si es de Perú, 0% si es exportación moderna"),
        ("Monto de IGV", "=B33*B34", "Soles (PEN)", "IGV local cobrado al cliente"),
        ("Total Facturado Proforma (Con IGV)", "=B33+B35", "Soles (PEN)", "Total final a facturar en el comprobante electrónico"),
        ("Detracción Retenida (12% SPOT)", "=IF(AND(B31=\"PE\", B36>700), B36*Configuración!$B$11, 0.0)", "Soles (PEN)", "12% retenido por cliente local y depositado en Banco de la Nación"),
        ("Pago a Cuenta RMT Mensual (1% SUNAT)", "=B33*Configuración!$B$13", "Soles (PEN)", "Impuesto mensual a pagar a cuenta (Régimen MYPE)"),
        ("Provisión RMT Impuesto Anual (10% RMT)", "=IF((B33-B25)>0, (B33-B25)*Configuración!$B$14, 0.0)", "Soles (PEN)", "Reserva del 10% sobre la utilidad para la declaración anual"),
        ("Comisión Pasarela / Fricción (PayPal/Ligo)", "=B33-B27", "Soles (PEN)", "Costo real descontado por pasarelas financieras"),
        ("Monto Neto Disponible (Flujo de Caja Real)", "=B36-B37-B38-B39-B40", "Soles (PEN)", "Liquidez neta disponible tras pagar impuestos y pasarelas")
    ]
    
    for idx, tax in enumerate(tax_data, 31):
        ws_calc.cell(row=idx, column=1, value=tax[0]).font = regular_font
        
        val_cell = ws_calc.cell(row=idx, column=2, value=tax[1])
        val_cell.font = bold_font
        val_cell.alignment = Alignment(horizontal="right")
        
        if tax[0] == "Tasa de IGV":
            val_cell.number_format = '0.0%'
        elif tax[0].startswith("Ubicación del Cliente") or tax[0].startswith("Canal de Pago"):
            val_cell.fill = warning_yellow_fill
            val_cell.alignment = Alignment(horizontal="center")
        else:
            val_cell.number_format = 'S/. #,##0.00'
            
        ws_calc.cell(row=idx, column=3, value=tax[2]).font = regular_font
        ws_calc.cell(row=idx, column=4, value=tax[3]).font = regular_font
        
        for c in range(1, 5):
            cell = ws_calc.cell(row=idx, column=c)
            cell.border = border_all
            if tax[0] == "Monto Neto Disponible (Flujo de Caja Real)":
                cell.fill = accent_green_fill
                cell.font = bold_font
                cell.border = border_total
            elif tax[0] in ["Precio Sugerido con Gross-Up", "Total Facturado Proforma (Con IGV)"]:
                cell.fill = light_blue_fill
                cell.font = bold_font

    # ----------------------------------------------------
    # SHEET 4: ACU DE SERVICIOS
    # ----------------------------------------------------
    ws_acu = wb.create_sheet(title="ACU de Servicios")
    ws_acu.views.sheetView[0].showGridLines = True
    
    ws_acu["A1"] = "ANÁLISIS DE COSTOS UNITARIOS (ACU) - SERVICIOS BASE"
    ws_acu["A1"].font = title_font
    ws_acu["A2"] = "Cálculos estructurados mediante el modelo TDABC para el portafolio de 13 servicios."
    ws_acu["A2"].font = subtitle_font
    
    headers_acu = [
        "ID", "Nombre del Servicio", "Categoría", "Hrs Alberto", "Hrs Junior", "Hrs QA", 
        "Materiales (PEN)", "MOD (PEN)", "MID (PEN)", "CIF/Overhead (PEN)", 
        "Break-Even (PEN)", "P. Seguridad (PEN)", "Venta Sugerida (PEN)"
    ]
    for col_idx, h in enumerate(headers_acu, 1):
        cell = ws_acu.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    services_acu = [
        ("C.1", "Migración Zapier a n8n", "Automatización", 3, 6, 0, 20.74),
        ("C.2.A", "Automatización Simple (3 Apps)", "Automatización", 1, 3, 0, 0.00),
        ("B.1", "Silhouette OS Lead Triage", "Automatización / IA", 10, 15, 3, 95.63),
        ("A.1", "LangGraph Swarms & Memory", "IA / Deep Tech", 25, 25, 5, 56.25),
        ("C.2.B", "WordPress Web / LMS / Shop", "Desarrollo Web", 3, 10, 1, 470.63),
        ("C.2.C", "Web Corporativa / Landing", "Desarrollo Web", 2, 6, 0, 150.00),
        ("B.2", "App Web React/TS (SPA)", "Desarrollo Web", 10, 20, 4, 76.88),
        ("RET.B", "Mantenimiento Básico (SLA)", "Mantenimiento", 1, 2, 0, 11.25),
        ("RET.A", "Mantenimiento IA/SaaS (Adv)", "Mantenimiento", 2, 3, 0, 37.50),
        ("A.2", "SaaS Completo Next.js / Stripe", "Software Corp", 25, 40, 8, 206.25),
        ("B.2.B", "Software Corp Custom (ERP/CRM)", "Software Corp", 20, 30, 6, 131.25),
        ("B.2.C", "Onboarding CRM HubSpot", "Implementación", 4, 6, 0, 0.00),
        ("CON.1", "Consultoría IA & Prompts", "Consultoría", 5, 0, 0, 0.00)
    ]
    
    for row_idx, s in enumerate(services_acu, 5):
        ws_acu.cell(row=row_idx, column=1, value=s[0]).font = regular_font
        ws_acu.cell(row=row_idx, column=2, value=s[1]).font = regular_font
        ws_acu.cell(row=row_idx, column=3, value=s[2]).font = regular_font
        
        # Hrs input
        for c in range(4, 7):
            cell = ws_acu.cell(row=row_idx, column=c, value=s[c-1])
            cell.font = regular_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'
            
        # Materiales input (col G)
        mat_cell = ws_acu.cell(row=row_idx, column=7, value=s[6])
        mat_cell.font = regular_font
        mat_cell.alignment = Alignment(horizontal="right")
        mat_cell.number_format = 'S/. #,##0.00'
        
        # Formulas
        r = row_idx
        ws_acu.cell(row=r, column=8, value=f"=(D{r}*Configuración!$B$5)+(E{r}*Configuración!$B$7)+(F{r}*Configuración!$B$8)")
        ws_acu.cell(row=r, column=9, value=f"=G{r}")
        if s[0] == "RET.B":
            ws_acu.cell(row=r, column=10, value="=60.00")
        elif s[0] == "RET.A":
            ws_acu.cell(row=r, column=10, value="=120.00")
        else:
            ws_acu.cell(row=r, column=10, value=f"=(D{r}+E{r}+F{r})*'Overhead y Capacidad'!$B$27")
            
        ws_acu.cell(row=r, column=11, value=f"=(H{r}+I{r}+J{r})*(1+0.085)")
        ws_acu.cell(row=r, column=12, value=f"=K{r}/(1-Configuración!$B$18)")
        ws_acu.cell(row=r, column=13, value=f"=K{r}/(1-Configuración!$B$19)")
        
        # Style calculated columns
        for col_idx in range(8, 14):
            cell = ws_acu.cell(row=r, column=col_idx)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = 'S/. #,##0.00'
            if col_idx == 11:
                cell.fill = light_blue_fill
            elif col_idx in [12, 13]:
                cell.fill = accent_green_fill
                
        for col_i in range(1, 14):
            ws_acu.cell(row=r, column=col_i).border = border_all

    # ----------------------------------------------------
    # SHEET 5: ACU DE ADICIONALES
    # ----------------------------------------------------
    ws_acux = wb.create_sheet(title="ACU de Adicionales")
    ws_acux.views.sheetView[0].showGridLines = True
    
    ws_acux["A1"] = "ANÁLISIS DE COSTOS UNITARIOS (ACU) - EXTRAS Y ADICIONALES"
    ws_acux["A2"] = "Cálculos de costos base y precios de seguridad para módulos y mejoras."
    ws_acux["A2"].font = subtitle_font
    
    # Headers
    for col_idx, h in enumerate(headers_acu, 1):
        cell = ws_acux.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    extras_acu = [
        ("EX-1", "Idioma Adicional (Multilenguaje)", "Módulo extra", 1, 2, 0, 7.50),
        ("EX-2", "Conexión a Pasarela Local", "Integración", 1, 4, 0, 0.00),
        ("EX-3", "Módulo de WhatsApp / Meta Extra", "Canal Extra", 1, 2, 0, 15.00),
        ("EX-4", "Flujo Extra de n8n", "Automatización", 1, 2, 0, 0.00),
        ("EX-5", "Optimización de Velocidad (WPO)", "Mejora", 1, 4, 0, 75.00),
        ("EX-6", "Panel de Reportes Analíticos con IA", "Módulo extra", 5, 10, 2, 37.50),
        ("EX-7", "Capacitación Técnica Extra", "Soporte", 1, 0, 0, 0.00),
        ("EX-8", "Dominio y SSL Setup", "Infraestructura", 0, 1, 0, 0.00)
    ]
    
    for row_idx, s in enumerate(extras_acu, 5):
        ws_acux.cell(row=row_idx, column=1, value=s[0]).font = regular_font
        ws_acux.cell(row=row_idx, column=2, value=s[1]).font = regular_font
        ws_acux.cell(row=row_idx, column=3, value=s[2]).font = regular_font
        
        # Hrs input
        for c in range(4, 7):
            cell = ws_acux.cell(row=row_idx, column=c, value=s[c-1])
            cell.font = regular_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'
            
        # Materiales input
        mat_cell = ws_acux.cell(row=row_idx, column=7, value=s[6])
        mat_cell.font = regular_font
        mat_cell.alignment = Alignment(horizontal="right")
        mat_cell.number_format = 'S/. #,##0.00'
        
        # Formulas
        r = row_idx
        ws_acux.cell(row=r, column=8, value=f"=(D{r}*Configuración!$B$5)+(E{r}*Configuración!$B$7)+(F{r}*Configuración!$B$8)")
        ws_acux.cell(row=r, column=9, value=f"=G{r}")
        ws_acux.cell(row=r, column=10, value=f"=(D{r}+E{r}+F{r})*'Overhead y Capacidad'!$B$27")
        ws_acux.cell(row=r, column=11, value=f"=(H{r}+I{r}+J{r})*(1+0.085)")
        ws_acux.cell(row=r, column=12, value=f"=K{r}/(1-Configuración!$B$18)")
        ws_acux.cell(row=r, column=13, value=f"=K{r}/(1-Configuración!$B$19)")
        
        # Style calculated columns
        for col_idx in range(8, 14):
            cell = ws_acux.cell(row=r, column=col_idx)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = 'S/. #,##0.00'
            if col_idx == 11:
                cell.fill = light_blue_fill
            elif col_idx in [12, 13]:
                cell.fill = accent_green_fill
                
        for col_i in range(1, 14):
            ws_acux.cell(row=r, column=col_i).border = border_all

    # ----------------------------------------------------
    # SHEET 6: CATÁLOGO COMERCIAL
    # ----------------------------------------------------
    ws_cat = wb.create_sheet(title="Catálogo Comercial")
    ws_cat.views.sheetView[0].showGridLines = True
    
    ws_cat["A1"] = "CATÁLOGO DE SERVICIOS Y TARIFARIO COMERCIAL"
    ws_cat["A1"].font = title_font
    ws_cat["A2"] = "Resumen consolidado para impresión o cotización directa a clientes (PEN)."
    ws_cat["A2"].font = subtitle_font
    
    # Table headers
    headers_cat = ["Código", "Servicio / Módulo", "Categoría", "Costo Break-Even", "Precio Seguridad", "Precio Sugerido", "Público Objetivo"]
    for col_idx, h in enumerate(headers_cat, 1):
        cell = ws_cat.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    cat_items = [
        # Servicio Base
        ("C.1", "='ACU de Servicios'!B5", "='ACU de Servicios'!C5", "='ACU de Servicios'!K5", "='ACU de Servicios'!L5", "='ACU de Servicios'!M5", "Público C - Pymes"),
        ("C.2.A", "='ACU de Servicios'!B6", "='ACU de Servicios'!C6", "='ACU de Servicios'!K6", "='ACU de Servicios'!L6", "='ACU de Servicios'!M6", "Público C - Pymes"),
        ("B.1", "='ACU de Servicios'!B7", "='ACU de Servicios'!C7", "='ACU de Servicios'!K7", "='ACU de Servicios'!L7", "='ACU de Servicios'!M7", "Público B - Agencias / Startups"),
        ("A.1", "='ACU de Servicios'!B8", "='ACU de Servicios'!C8", "='ACU de Servicios'!K8", "='ACU de Servicios'!L8", "='ACU de Servicios'!M8", "Público A - Scale-ups / Corp"),
        ("C.2.B", "='ACU de Servicios'!B9", "='ACU de Servicios'!C9", "='ACU de Servicios'!K9", "='ACU de Servicios'!L9", "='ACU de Servicios'!M9", "Público C - E-Commerce / LMS"),
        ("C.2.C", "='ACU de Servicios'!B10", "='ACU de Servicios'!C10", "='ACU de Servicios'!K10", "='ACU de Servicios'!L10", "='ACU de Servicios'!M10", "Público C - Landing / Pymes"),
        ("B.2", "='ACU de Servicios'!B11", "='ACU de Servicios'!C11", "='ACU de Servicios'!K11", "='ACU de Servicios'!L11", "='ACU de Servicios'!M11", "Público B - App Web React"),
        ("RET.B", "='ACU de Servicios'!B12", "='ACU de Servicios'!C12", "='ACU de Servicios'!K12", "='ACU de Servicios'!L12", "='ACU de Servicios'!M12", "Público C/B - Soporte SLA mensual"),
        ("RET.A", "='ACU de Servicios'!B13", "='ACU de Servicios'!C13", "='ACU de Servicios'!K13", "='ACU de Servicios'!L13", "='ACU de Servicios'!M13", "Público B/A - Soporte IA mensual"),
        ("A.2", "='ACU de Servicios'!B14", "='ACU de Servicios'!C14", "='ACU de Servicios'!K14", "='ACU de Servicios'!L14", "='ACU de Servicios'!M14", "Público A - Plataformas SaaS Next.js"),
        ("B.2.B", "='ACU de Servicios'!B15", "='ACU de Servicios'!C15", "='ACU de Servicios'!K15", "='ACU de Servicios'!L15", "='ACU de Servicios'!M15", "Público B - Software Corp ERP"),
        ("B.2.C", "='ACU de Servicios'!B16", "='ACU de Servicios'!C16", "='ACU de Servicios'!K16", "='ACU de Servicios'!L16", "='ACU de Servicios'!M16", "Público B - Onboarding CRM"),
        ("CON.1", "='ACU de Servicios'!B17", "='ACU de Servicios'!C17", "='ACU de Servicios'!K17", "='ACU de Servicios'!L17", "='ACU de Servicios'!M17", "Público B/A - Consultoría IA"),
        
        # Extras
        ("EX-1", "='ACU de Adicionales'!B5", "='ACU de Adicionales'!C5", "='ACU de Adicionales'!K5", "='ACU de Adicionales'!L5", "='ACU de Adicionales'!M5", "Módulo Adicional Multilenguaje"),
        ("EX-2", "='ACU de Adicionales'!B6", "='ACU de Adicionales'!C6", "='ACU de Adicionales'!K6", "='ACU de Adicionales'!L6", "='ACU de Adicionales'!M6", "Pasarela Local (Culqi/Niubiz)"),
        ("EX-3", "='ACU de Adicionales'!B7", "='ACU de Adicionales'!C7", "='ACU de Adicionales'!K7", "='ACU de Adicionales'!L7", "='ACU de Adicionales'!M7", "Meta WhatsApp API Módulo Extra"),
        ("EX-4", "='ACU de Adicionales'!B8", "='ACU de Adicionales'!C8", "='ACU de Adicionales'!K8", "='ACU de Adicionales'!L8", "='ACU de Adicionales'!M8", "Flujo Automatización Extra n8n"),
        ("EX-5", "='ACU de Adicionales'!B9", "='ACU de Adicionales'!C9", "='ACU de Adicionales'!K9", "='ACU de Adicionales'!L9", "='ACU de Adicionales'!M9", "Optimización de Carga WPO Web"),
        ("EX-6", "='ACU de Adicionales'!B10", "='ACU de Adicionales'!C10", "='ACU de Adicionales'!K10", "='ACU de Adicionales'!L10", "='ACU de Adicionales'!M10", "Dashboard Analítico en Silhouette OS"),
        ("EX-7", "='ACU de Adicionales'!B11", "='ACU de Adicionales'!C11", "='ACU de Adicionales'!K11", "='ACU de Adicionales'!L11", "='ACU de Adicionales'!M11", "Capacitación a Personal Técnico / hora"),
        ("EX-8", "='ACU de Adicionales'!B12", "='ACU de Adicionales'!C12", "='ACU de Adicionales'!K12", "='ACU de Adicionales'!L12", "='ACU de Adicionales'!M12", "Registro de Dominio y SSL Setup")
    ]
    
    for row_idx, item in enumerate(cat_items, 5):
        ws_cat.cell(row=row_idx, column=1, value=item[0]).font = regular_font
        ws_cat.cell(row=row_idx, column=2, value=item[1]).font = regular_font
        ws_cat.cell(row=row_idx, column=3, value=item[2]).font = regular_font
        
        # Prices
        for c in range(4, 7):
            cell = ws_cat.cell(row=row_idx, column=c, value=item[c-1])
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = 'S/. #,##0.00'
            if c == 4:
                cell.fill = light_blue_fill
            else:
                cell.fill = accent_green_fill
                
        ws_cat.cell(row=row_idx, column=7, value=item[6]).font = regular_font
        
        for col_i in range(1, 8):
            ws_cat.cell(row=row_idx, column=col_i).border = border_all
            if item[0].startswith("EX"):
                ws_cat.cell(row=row_idx, column=col_i).fill = orange_fill

    # ----------------------------------------------------
    # SHEET 7: SIMULADOR DE FLUJO DE CAJA (ANUAL)
    # ----------------------------------------------------
    ws_flow = wb.create_sheet(title="Simulador de Caja")
    ws_flow.views.sheetView[0].showGridLines = True
    
    ws_flow["A1"] = "SIMULADOR DE FLUJO DE CAJA ANUAL PROYECTADO"
    ws_flow["A1"].font = title_font
    ws_flow["A2"] = "Modifica el número de proyectos vendidos al mes (celdas amarillas) para proyectar tu flujo de caja real anual bajo el RMT."
    ws_flow["A2"].font = subtitle_font
    
    headers_flow = [
        "Mes", "Proy. C (C.1)", "Proy. B (B.1)", "Proy. A (A.1)", "Retainers (RET.B)",
        "Ingresos Brutos (S/.)", "Overhead Fijo (S/.)", "Nómina Apoyo (S/.)", 
        "Impuesto RMT (1%) (S/.)", "Comisiones (8.5%) (S/.)", "Provisión Renta (10%) (S/.)", "Flujo Disponible (S/.)"
    ]
    for col_idx, h in enumerate(headers_flow, 1):
        cell = ws_flow.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    for idx, month in enumerate(months, 5):
        r = idx
        ws_flow.cell(row=r, column=1, value=month).font = regular_font
        
        # Default yellow inputs for quantity of sales per month
        for col in range(2, 6):
            cell = ws_flow.cell(row=r, column=col)
            cell.font = bold_font
            cell.fill = warning_yellow_fill
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0'
            
        # Pre-load reasonable default sales assumptions
        ws_flow.cell(row=r, column=2, value=1 if r % 2 == 0 else 0)  # C.1: 0.5 per month average
        ws_flow.cell(row=r, column=3, value=1 if r % 3 == 0 else 0)  # B.1: 0.33 per month
        ws_flow.cell(row=r, column=4, value=1 if r == 8 else 0)      # A.1: 1 project per year
        ws_flow.cell(row=r, column=5, value=2 if r < 8 else 4)       # Retainers: 2 scaling to 4
        
        # Formulas
        # F: Ingresos Brutos = Proy_C * Sugerido_C1 + Proy_B * Sugerido_B1 + Proy_A * Sugerido_A1 + Ret * Sugerido_RET
        ws_flow.cell(row=r, column=6, value=f"=(B{r}*'Catálogo Comercial'!$F$5)+(C{r}*'Catálogo Comercial'!$F$7)+(D{r}*'Catálogo Comercial'!$F$8)+(E{r}*'Catálogo Comercial'!$F$12)")
        
        # G: Overhead Fijo = Overhead y Capacidad!$B$26
        ws_flow.cell(row=r, column=7, value="='Overhead y Capacidad'!$B$26")
        
        # H: Nómina Apoyo (Junior & QA hours * rates for each project type)
        # C.1 Support = (6*30 + 0*22.5) = 180.00
        # B.1 Support = (15*30 + 3*22.5) = 517.50
        # A.1 Support = (25*30 + 5*22.5) = 862.50
        # RET.B Support = (2*30) = 60.00
        ws_flow.cell(row=r, column=8, value=f"=(B{r}*180.00)+(C{r}*517.50)+(D{r}*862.50)+(E{r}*60.00)")
        
        # I: Impuesto RMT Pago a cuenta (1.0% del bruto)
        ws_flow.cell(row=r, column=9, value=f"=F{r}*Configuración!$B$13")
        
        # J: Comisiones Pasarela (PayPal 8.5% average)
        ws_flow.cell(row=r, column=10, value=f"=F{r}*0.085")
        
        # K: Provisión Renta Anual (10% de utilidad neta estimada)
        # Utilidad = Ingresos - Gastos_Fijos - Nómina - Comisiones. Provisión = 10% of utility if positive.
        ws_flow.cell(row=r, column=11, value=f"=IF((F{r}-G{r}-H{r}-J{r})>0, (F{r}-G{r}-H{r}-J{r})*Configuración!$B$14, 0.0)")
        
        # L: Flujo Disponible = Ingresos - Gastos_Fijos - Nómina - Impuesto_RMT - Comisiones - Provisión_Renta
        ws_flow.cell(row=r, column=12, value=f"=F{r}-G{r}-H{r}-I{r}-J{r}-K{r}")
        
        # Formatting calculated cells
        for c in range(6, 13):
            cell = ws_flow.cell(row=r, column=c)
            cell.font = bold_font if c in [6, 12] else regular_font
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = 'S/. #,##0.00'
            if c == 12:
                cell.fill = accent_green_fill
            elif c == 6:
                cell.fill = light_blue_fill
                
        for col_i in range(1, 13):
            ws_flow.cell(row=r, column=col_i).border = border_all
            
    # Totals Row at row 17
    ws_flow.cell(row=17, column=1, value="TOTAL ANUAL").font = bold_font
    ws_flow.cell(row=17, column=1).alignment = Alignment(horizontal="center")
    
    for col in range(2, 6):
        cell = ws_flow.cell(row=17, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}16)")
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = '#,##0'
        
    for col in range(6, 13):
        cell = ws_flow.cell(row=17, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}16)")
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = 'S/. #,##0.00'
        if col == 12:
            cell.fill = accent_green_fill
        elif col == 6:
            cell.fill = light_blue_fill
            
    for col_i in range(1, 13):
        ws_flow.cell(row=17, column=col_i).border = border_total

    # Auto-adjust column widths
    for ws in [ws_config, ws_oh, ws_calc, ws_acu, ws_acux, ws_cat, ws_flow]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    max_len = max(max_len, 14)
                else:
                    max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save
    file_path = "Calculadora_de_Costos_y_Precios.xlsx"
    wb.save(file_path)
    print(f"Workbook saved to {file_path}")

if __name__ == "__main__":
    create_excel_workbook()
